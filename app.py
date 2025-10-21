import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from pathlib import Path
import re


# -------------------------------------------------------------------------
#                             Bancos de México 
#  ------------------------------------------------------------------------

# 1. Diccionario de reglas por banco

reglas_bancos_mx = {
    "BBVA": {
        "columnas": {
            "cuenta": 0,
            "fecha": 1,
            "fecha_ope": 1,
            "concepto": [2,3,4],
            "cargo": 5,
            "abono": 6,    
        },
        "tipo_importe": "abono_cargo" ,
        "separador_miles": ".",
        "separador_decimales": ","
    },

        "Banorte": {
        "columnas": {
            "cuenta": 0,
            "fecha": 2,
            "fecha_ope": 1,
            "concepto": 11,
            "cargo": 8,
            "abono": 7,    
        },
        "tipo_importe": "" ,
    },

        "Edenred": {
        "columnas": {
            "cuenta": 0,
            "fecha": 0,
            "fecha_ope": 0,
            "concepto": 2,
            "cargo": 6,
            "abono": 5,
            "ref 1": 3,
        },
        "tipo_importe": "" ,
    }
}

# 2. Función para parsear múltiples formatos de fecha

formatos_fecha_mx = [
    "%d.%m.%y",
    "%d.%m.%Y",
    "%d/%m/%Y",
    "%Y%m%d",
    "%Y-%m-%d",
    "%d-%m-%Y"
]

def parsear_fecha_multiple_mx(valor):
    """Intenta convertir un valor a fecha probando varios formatos."""
    if pd.isna(valor):              # Verifica si el valor está vacío o es nulo
        return pd.NaT               # Si el valor es nulo (por ejemplo, está vacío o es NaN), devuelve pd.NaT (que significa "Not a Time", es decir, no hay fecha).
    valor = str(valor).strip()      # Convierte el valor a texto y elimina espacios en blanco al inicio y al final
    
    
    # Para los casos donde la fecha trae más información como la hora. Por ej: 1/08/2025  12:27:45 p. m.
    if " " in valor:
        valor = valor.split(" ")[0]
    
    for formato in formatos_fecha_mx:  # Intenta convertir el valor a fecha usando varios formatos
        try:
            return pd.to_datetime(valor, format=formato, errors="raise")
        except:
            continue
    return pd.NaT


# 4. Función calcular importe


def calcular_importe(df, reglas, banco=None):
    """
    Calcula el importe según las reglas del banco.
    - reglas["columnas"]["abono"]: índice de columna de abonos
    - reglas["columnas"]["cargo"]: índice de columna de cargos
    - reglas.get("tipo_importe"): 'abono_cargo' o 'cargo_abono'
    """

    columnas = reglas['columnas']
    tipo = reglas.get("tipo_importe", "abono_cargo")

    # Extraer columnas
    abono_col = df.iloc[:, columnas['abono']].astype(str).str.strip()
    cargo_col = df.iloc[:, columnas['cargo']].astype(str).str.strip()

    # 🚨 Limpieza especial solo para Banorte
    
    def limpiar_columna(serie):
        return (
            serie
            .str.replace(r'[^0-9,.-]', '', regex=True)  # quita $ y otros
            .str.replace(',', '', regex=False)          # elimina comas de miles
            .replace('', '0')                           # si queda vacío → 0
        )
    if banco in ["Banorte", "Edenred"]:
        abono_col = limpiar_columna(abono_col)
        cargo_col = limpiar_columna(cargo_col) 
    
       # Convertir a numérico
    abono = pd.to_numeric(abono_col, errors="coerce").fillna(0)
    cargo = pd.to_numeric(cargo_col, errors="coerce").fillna(0)

    # Retornar según el tipo de cálculo
    return abono - cargo if tipo == "abono_cargo" else -cargo + abono

# 5. Función genérica de transformación para bancos de México

def transformar_extracto_mx(df,banco, archivo=None):
    
    if banco not in reglas_bancos_mx:
        raise ValueError(f"No hay reglas definidas para el banco '{banco}'")

    reglas = reglas_bancos_mx.get(banco)
    columnas = reglas['columnas']
    
    df_out_mx = df.copy()

    # Mapear columnas según reglas

    # ✅ Columnas: fechas
    df_out_mx['fecha_ope'] = (
        df.iloc[:, columnas['fecha_ope']]
        .astype(str)
        .str.strip()
        .apply(parsear_fecha_multiple_mx)
    )
    df_out_mx['fecha_ope'] = df_out_mx['fecha_ope'].dt.strftime("%d/%m/%Y")

    df_out_mx['fecha'] = (
        df.iloc[:, columnas['fecha']]
        .astype(str)
        .str.strip()
        .apply (parsear_fecha_multiple_mx)
    )
    df_out_mx['fecha'] = df_out_mx['fecha'].dt.strftime("%d/%m/%Y")

 # ✅ Columnas opcionales
    opcionales_mx = {
        'ref 1': lambda s: s.astype(str).str.lstrip('0').str.upper(),
        'ref 2': lambda s: s.astype(str).str.lstrip('0').str.upper()
    }

    for col, func in opcionales_mx.items():
        if col in columnas:
            df_out_mx[col] = func(df.iloc[:, columnas[col]])
        else:
            df_out_mx[col] = ""

# ✅ Columna: concepto
   
   # Concatenar varias columnas si es necesario o tomar información de una sola
    concepto_cols = columnas['concepto']
    if isinstance(concepto_cols, list):
        # concatenar columnas en orden   
        df_out_mx['concepto'] =(
            df.iloc[:, concepto_cols] # seleccionamos varias columnas (ej. [1,2,3])
            .astype(str)              # convertimos a texto
            .apply(lambda fila: ' '.join(fila).strip(), axis=1) # concatenamos y eliminamos espacios en blanco al inicio y al final
            )
    else:
        # una sola columna
        df_out_mx["concepto"] = df.iloc[:, concepto_cols].astype(str)

# ✅ Columna: importe

    df_out_mx['importe'] = calcular_importe(df, reglas,banco=banco)

# ✅ Columna: cuenta

    df_out_mx['cuenta'] = df.iloc[:, columnas['cuenta']].astype(str).str.strip()
    
    # Formato especial según banco

    if banco == "Banorte":
        df_out_mx['cuenta'] = "BANORTE " + df_out_mx['cuenta'].str[-4:]
    elif banco == "Edenred":
        df_out_mx['cuenta'] = "EDENRED"
                                                               
    df_final_mx = df_out_mx[['cuenta','fecha', 'fecha_ope', 'concepto', 'importe', 'ref 1', 'ref 2']]
    return df_final_mx


# -------------------------- Interfaz Streamlit --------------------------
st.title("Transformador de Extractos")

st.subheader("🏦 Bancos de México")

# Seleccionar banco
banco_seleccionado_mx = st.selectbox(
    "Selecciona el banco",
    options=list(reglas_bancos_mx.keys())
)

# Subir archivo
archivos_mx = st.file_uploader(
    "📂 Carga tus extractos",
    type=["txt","csv", "xlsx"],
    accept_multiple_files=True,
    key="uploader_mx")

# Lista para guardar los resultados de cada archivo transformado

dfs_transformados_mx = []
archivos_cargados_mx = []

if archivos_mx is not None: # Verifica si hay archivos cargados
    for archivo in archivos_mx:
        try:

            nombre = archivo.name.lower()

            # Detectar tipo de archivo por extensión
            if nombre.endswith(".txt"):
                df = pd.read_csv(archivo, sep=';', decimal=",", encoding='latin1', header=None)
            elif nombre.endswith(".csv"):
                df = pd.read_csv(archivo, sep=",", decimal=".", encoding="latin1", header=None, skiprows=1)
            elif nombre.endswith(".xlsx"):
                    if banco_seleccionado_mx == "BBVA":
                        df = pd.read_excel(archivo, header=None, skiprows=2)  # Omitir 2 primeras filas
                    else:
                        df = pd.read_excel(archivo, header=None, skiprows=1)
            else:
                st.warning(f"Formato no compatible: {archivo.name}")
                continue    

            # Transformar archivo
            df_transformado_mx = transformar_extracto_mx(df, banco=banco_seleccionado_mx, archivo=archivo)
            dfs_transformados_mx.append(df_transformado_mx)
            archivos_cargados_mx.append(f"✅ {archivo.name}")

        except Exception as e: 
            archivos_cargados_mx.append(f"❌ {archivo.name} (Error: {e})")     
    
    # Mostrar resumen en un expander
    with st.expander("Ver archivos cargados y estado"):
        for estado in archivos_cargados_mx:
            st.write(estado)

    if dfs_transformados_mx:
        df_transformado_mx = pd.concat(dfs_transformados_mx, ignore_index=True)

        st.success("✅ Archivos procesados y consolidados correctamente") 

        # Mostrar vista previa del consolidado
        st.subheader(f"Vista previa:")
        st.dataframe(df_transformado_mx.head(5))        


        # Descargar archivo en Excel
        buffer = BytesIO()
        df_transformado_mx.to_excel(buffer, index=False, engine='openpyxl')
        buffer.seek(0)

        # Aplicar formato con openpyxl
        wb = load_workbook(buffer)
        ws = wb.active

        if "importe" in df_transformado_mx.columns:
            # Ubicar la columna "importe" (posición 12 en df_final)
            col_importe = df_transformado_mx.columns.get_loc('importe') + 1
            for row in ws.iter_rows(min_row=2, min_col=col_importe, max_col=col_importe):
                for cell in row:
                    cell.number_format = '#,##0.00;[Red]-#,##0.00' # miles con "." y decimales con "," y negativos en rojo        

            # Guardar en nuevo buffer
            buffer_final_mx = BytesIO()
            wb.save(buffer_final_mx)
            buffer_final_mx.seek(0)

            st.download_button(
                label="📥 Descargar extractos consolidados",
                data=buffer_final_mx,
                file_name=f"{banco_seleccionado_mx} - extractos_transformados.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )  


# -----------------------------------------------------------------------
#                             Bancos de Colombia 
#  ----------------------------------------------------------------------

# 1. Diccionarios
# 1.1 Diccionario de códigos y su descripción

codigos_dict = {
    "304": "Pago Tarjeta de Credito Banco Bogota Internet o Banca Movil",
    "569": "Pago por Internet Corporativo",
    "715": "Pago impuesto Nal x Int. Corporativo",
    "717": "Recarga Tarjeta Efectiva",
    "853": "Abono compra divisas",
    "670": "Cargo Pago Electronico Planilla Unica No. 000000000000000000000047664212 Nit 0009998600669427",
    "159": "Comision dispersion de pago de proveedores-Otros",
    "183": "Comision transaccion AVAL",
    "502": "Cargo comision consignacion",
    "595": "Comision dispersion pago de nomina",
    "854": "Cargo internacional 20176321000142 Registro Giros",
    "997": "CargoXpor Comision por Cheque de Gerencia Otras oficinas",
    "GT09": "Gravamen Movimientos Financieros",
    "GT10": "Abono ajuste gravamen mov. financieros",
    "GT26": "Cargo IVA",
    "161": "Abono devolucion dispersion pago de proveedores - otros",
    "594": "Abono devolucion dispersion pago de nomina",
    "12": "Pago cheque canje",
    "20": "Consignacion en oficina en cheque",
    "21": "Consignacion nacional en cheque",
    "26": "Transferencia de cuentas del Banco de Bogota",
    "60": "Abono transferencia por canal electrónico",
    "130": "Pago cheque a terceros Chequera No. 3001010628",
    "139": "Pago cheque en oficina Chequera No. 3001010628",
    "160": "Abono dispersion pago a proveedores",
    "164": "Transf",
    "201": "Cargo Dispersion Pago de Proveedores/Otros",
    "216": "Cargo transferencia por Internet o Banca Movil o Swift",
    "220": "Abono transferencia por internet o banca movil",
    "222": "Consignacion local en cheque",
    "276": "Consignacion en cheque",
    "397": "Abono a cuenta en oficina",
    "591": "Cargo Dispersion Pago de Nomina",
    "593": "Cr Ach",
    "611": "Abono por Deposito en Corresponsal de cliente",
    "659": "Cargo Pago de Cartera",
    "929": "Abono por deposito en cajero automatico",
    "995": "CargoXpor Compra de Cheque de Gerencia",
    "2180": "NA",
    "658": "Pago automatico cuota de credito",
    "524": "Cargo reversion recaudo ACH por RECFON",
    "679": "Retefuente remuneracion especial",
    "678": "Interes remuneracion especial",
    "509": "TIMBRE CHEQUERA",
    "508": "IVA CHEQUERA",
    "110": "COMPRA CHEQUERA",
    "GT01": "Intereses ganados",
    "665": "Carga giro empresarial",
    "219": "Pago de servicio o comparendo por canales electronicos",
    "482": "Pago servicio publico por internet o banca movil",
    "106": "Abono por dispersion de fondos por ATH",
    "91": "Abono por recaudos con comprobante",
    "86": "Cargo comision recaudos con comprobante",
    "GT08": "Cargo IVA",
    "570": "Abono recaudo pago electronico ACH",
    "90": "Pago automatico tarjeta de credito",
    "918": "Reversion comision",
    "GT06": "Retencion en la Fuente sobre intereses",
    "GT22": "Intereses por sobregiro",
    "221": "Abono transferencia AVAL por internet o banca movi",
    "644": "Comision PSE",
    "687": "Abono cancelacion CDT en proceso de prescripcion o prescrito",
    "394": "Cargo a cuenta en oficina",
    "938": "Devolucion IVA por ajuste a una comision",
}

# 1.2 Diccionario de cuentas y su ID

cuentas_bancos = {
    "291252245": 1,
    "223589391": 2,
    "040-000016-02": 3,
    "040-000068-06": 4,
    "040-000054-70": 5,
    "040-000038-64": 6,
    "040-000077-86": 7,
    "040-000045-62": 9,
    "4851-0000-3964": 12,
    "4851-6999-6280": 13,
    "IRIS 100598509191": 14,
    "FIC # 8287-1": 15,
    "FIC # 8287-3": 16,
    "FIC # 8287-4": 17,
    "2570": 18,
    "2580": 19,
    "0011-8": 20,
    "0164-0": 21,
    "BBVA 0016": 23,
    "171-2": 25
}

# 1.3 Diccionario de reglas por banco 

reglas_bancos = {
    "Banco de Bogotá": {
        "columnas": {
            "cuenta": 1,
            "fecha_ope": 3,
            "fecha": 13,
            "numero": 6,
            "it": 9,
            "importe": 10,
            "nit": 16,
            "nid": 18,
            "referencia": 21
        },
        "separador_miles": ".",
        "separador_decimales": ",",
        "codigo_tipo_transaccion": codigos_dict,
        "id": cuentas_bancos
    },

    "Bancolombia": {
        "columnas": {
            "cuenta": 0,
            "fecha_ope": 3,
            "fecha": 3,
            "numero": 6,
            "tipo_transaccion": 7,
            "importe": 5            
        },
        "separador_miles": ".",
        "separador_decimales": ",",
        "id": cuentas_bancos
    },

    "Davivienda": {
        "columnas": {
            "fecha_ope": 0,
            "fecha": 0,
            "numero": 6,
            "tipo_transaccion": 7,
            "importe": 8,
            "referencia": 2
        },
        "separador_miles": ",",
        "separador_decimales": ".",
        "id": cuentas_bancos
    } 
}


# 2. Función de limpieza de NIT

# Aplicar la regla de los 10 dígitos
def limpiar_nit(valor):
    if not valor or not isinstance(valor, str):  
        return valor
    if re.fullmatch(r"\d{10}", valor):    # si tiene exactamente 10 dígitos
        if 800 <= int(valor[:3]) <= 999:  # si empieza entre 800 y 999
            return valor[:-1]             # quitamos el último dígito
    return valor                          # si no cumple, lo dejamos igual


# 3. Función para parsear múltiples formatos de fecha

formatos_fecha = [
    "%d.%m.%y",
    "%d.%m.%Y",
    "%d/%m/%Y",
    "%Y%m%d",
    "%Y-%m-%d",
    "%d-%m-%Y"
]

def parsear_fecha_multiple(valor):
    """Intenta convertir un valor a fecha probando varios formatos."""
    if pd.isna(valor):              # Verifica si el valor está vacío o es nulo
        return pd.NaT               # Si el valor es nulo (por ejemplo, está vacío o es NaN), devuelve pd.NaT (que significa "Not a Time", es decir, no hay fecha).
    valor = str(valor).strip()      # Convierte el valor a texto y elimina espacios en blanco al inicio y al final
    for formato in formatos_fecha:  # Intenta convertir el valor a fecha usando varios formatos
        try:
            return pd.to_datetime(valor, format=formato, errors="raise")
        except:
            continue
    return pd.NaT

# 4. Función genérica de transformación para bancos de Colombia

def transformar_extracto(df,banco, archivo=None):
    """"
    df: DataFrame leido del archivo
    banco: nombre del banco en reglas_bancos
    archivo: objeto uploaded file de Streamlit (para extraer nombre)
    """

    if banco not in reglas_bancos:
        raise ValueError(f"No hay reglas definidas para el banco '{banco}'")

    reglas = reglas_bancos.get(banco)
    columnas = reglas['columnas']

    df_out = df.copy()
    
    # Mapear columnas según reglas

        # ✅ Columna: número
    df_out['numero'] = df.iloc[:, columnas ['numero']].astype(str).str.lstrip('0')

        
        # ✅ Columna: tipo_transaccion
    if 'tipo_transaccion' in columnas:
            # Si existe la columna, solo muestra el valor tal cual (como texto)
        df_out['tipo_transaccion'] = df.iloc[:, columnas ['tipo_transaccion']].astype(str)
    else: 
            # Si no existe la columna, busca el código en el diccionario, si no existe muestra 'Desconocido'
        df_out['tipo_transaccion'] = df_out['numero'].astype(str).map(codigos_dict).fillna('Desconocido')

        
        # ✅ Columnas: fechas
    df_out['fecha_ope'] = (
        df.iloc[:, columnas['fecha_ope']]
        .astype(str)
        .str.strip()
        .apply(parsear_fecha_multiple)
    )
    df_out['fecha_ope'] = df_out['fecha_ope'].dt.strftime("%d/%m/%Y")

    df_out['fecha'] = (
        df.iloc[:, columnas['fecha']]
        .astype(str)
        .str.strip()
        .apply (parsear_fecha_multiple)
    )
    df_out['fecha'] = df_out['fecha'].dt.strftime("%d/%m/%Y")

    df_out['día'] = pd.to_datetime(df_out['fecha_ope'], format="%d/%m/%Y", errors="coerce").dt.day  


        # ✅ Importe como número (float)
    df_out['importe'] = pd.to_numeric(df.iloc[:, columnas['importe']],errors="coerce").fillna(0)
   

        # ✅ Columnas opcionales
    opcionales = {
        'nit': lambda s: 
        s.astype(str)                           # aseguramos string
        .str.upper()                            # estandarizamos mayúsculas
        .str.replace(r"[A-Z]", "", regex=True)  # quitamos cualquier letra
        .str.strip()                            # quitamos espacios en blanco    
        .str.lstrip('0')                        # quitamos ceros a la izquierda 
        .apply(limpiar_nit),                    # aplicamos la regla de los 10 dígitos
        
        'it': lambda s: s.astype(str),
        'nid': lambda s: s.astype(str).str.lstrip('0'),
        'referencia': lambda s: s.astype(str).str.lstrip('0').str.upper()
    }

    for col, func in opcionales.items():
        if col in columnas:
            df_out[col] = func(df.iloc[:, columnas[col]])
        else:
            df_out[col] = ""


        # ✅ Columnas vacías obligatorias para mantener estructura
    df_out['i'] = ""
    df_out['descripcion'] = ""
    df_out['provisional'] = ""


        # ✅ Otros mapeos: Caso especial Davivienda-> Columna 'cuenta' se alimenta del nombre de archivo
    if banco == "Davivienda":
        if archivo is not None:
            nombre_archivo = archivo.name # Obtener el nombre del archivo cargado            
                    
                # limpiar el nombre para quitar caracteres raros y extraer el nombre sin extensión
            numero_cuenta =  re.sub(r'[^A-Za-z0-9_\-]', '',Path(nombre_archivo).stem) 
            df_out['cuenta'] = numero_cuenta
        else:
                # Si por alguna razón no tiene nombre_archivo, deja vacío
                df_out['cuenta'] = ""

                # Columna 'importe': según referencia (CREDITO -> positivo) (DEBITO -> negativo)
        referencia_davivienda = df_out['referencia'].fillna("").astype(str).str.upper()
        df_out.loc[referencia_davivienda.str.contains("DEBITO"), 'importe'] *= -1
              
                # mapear id por cuenta (si la cuenta está en el diccionario)
        df_out['id'] = df_out['cuenta'].map(cuentas_bancos).fillna('Desconocido')        
                
    else:
                # Otros bancos: cuenta viene de columna indicada en reglas
                # Protegemos el acceso por si falta la clave 'cuenta' en reglas
                
        if 'cuenta' in columnas:        
                df_out['cuenta'] = df.iloc[:, columnas['cuenta']].astype(str)
        else:
                df_out['cuenta'] = ""
    df_out['id'] = df_out['cuenta'].astype(str).map(cuentas_bancos).fillna('Desconocido')  


    # ✅ Estructura final
    df_final = df_out[['id', 'cuenta', 'fecha_ope', 'fecha', 'día', 'numero', 'tipo_transaccion', 'i', 'descripcion', 'it', 'provisional', 'importe','nit','nid', 'referencia']]
    return df_final

# -------------------------- Interfaz Streamlit --------------------------

st.subheader("🏦 Bancos de Colombia")

# Seleccionar banco
banco_seleccionado = st.selectbox(
    "Selecciona el banco",
    options=list(reglas_bancos.keys())
)
# Subir archivo
archivos = st.file_uploader(
    "📂 Carga tus extractos",
    type=["txt","csv", "xlsx"],
    accept_multiple_files=True,
    key="uploader_co")


# Lista para guardar los resultados de cada archivo transformado

dfs_transformados = []
archivos_cargados = []

if archivos is not None: # Verifica si hay archivos cargados
    for archivo in archivos:
        try:

            nombre = archivo.name.lower()

            # Detectar tipo de archivo por extensión
            if nombre.endswith(".txt"):
                df = pd.read_csv(archivo, sep=';', decimal=",", encoding='latin1', header=None)
            elif nombre.endswith(".csv"):
                df = pd.read_csv(archivo, sep=",", decimal=".", encoding="latin1", header=None)
            elif nombre.endswith(".xlsx"):
                    if banco_seleccionado == "Davivienda":
                        df = pd.read_excel(archivo, header=None, skiprows=3)  # ⬅️ Omitir 3 primeras filas
                    else:
                        df = pd.read_excel(archivo, header=None)
            else:
                st.warning(f"Formato no compatible: {archivo.name}")
                continue    

                  
            # Transformar archivo
            df_transformado = transformar_extracto(df, banco=banco_seleccionado, archivo=archivo)
            dfs_transformados.append(df_transformado)
            archivos_cargados.append(f"✅ {archivo.name}")

        except Exception as e: 
            archivos_cargados.append(f"❌ {archivo.name} (Error: {e})")     
    
    # Mostrar resumen en un expander
    with st.expander("Ver archivos cargados y estado"):
        for estado in archivos_cargados:
            st.write(estado)

    if dfs_transformados:
        df_transformado = pd.concat(dfs_transformados, ignore_index=True)

        st.success("✅ Archivos procesados y consolidados correctamente") 

        # Ordenar por id de forma ascendente
        df_transformado = df_transformado.sort_values(by="id", ascending=True).reset_index(drop=True)

        # Mostrar vista previa del consolidado
        st.subheader(f"Vista previa:")
        st.dataframe(df_transformado.head(5))   
            
        # Descargar archivo en Excel
        buffer = BytesIO()
        df_transformado.to_excel(buffer, index=False, engine='openpyxl')
        buffer.seek(0)

        # Aplicar formato con openpyxl
        wb = load_workbook(buffer)
        ws = wb.active        

            
        if "importe" in df_transformado.columns:
            # Ubicar la columna "importe" (posición 12 en df_final)
            col_importe = df_transformado.columns.get_loc('importe') + 1
            for row in ws.iter_rows(min_row=2, min_col=col_importe, max_col=col_importe):
                for cell in row:
                    cell.number_format = '#,##0.00;[Red]-#,##0.00' # miles con "." y decimales con "," y negativos en rojo        

            # Guardar en nuevo buffer
            buffer_final = BytesIO()
            wb.save(buffer_final)
            buffer_final.seek(0)

            st.download_button(
                label="📥 Descargar extractos consolidados",
                data=buffer_final,
                file_name=f"{banco_seleccionado} - extractos_transformados.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


# -----------------------------------------------------------------------
#                             Conciliación bancaria
#  ----------------------------------------------------------------------

# 1. Reglas de trasnformación segun el tipo de fuente 

reglas_fuente = {
        "Extracto": {
            "columnas": {
                "Cuenta": 0,
                "Fecha": 2,
                "Descripcion": 3,
                "Importe": 4,
                "Ref 1": 5,
                "Ref 2": 6,
            },

        },

            "Auxiliar": {
                "columnas": {
                    "Cuenta":8,
                    "Fecha": 4,
                    "Asiento": 5,
                    "Descripcion": 10,
                    "Importe": 15,
                    "Nid": 18,
                    "IT Tercero": 16,
                    "Tercero": 17,
                    "Responsable": 29,
                    "Fecha_hora modificacion": 31

        },
    }

}

# 2. Función para parsear múltiples formatos de fecha

formatos_fecha = [
    "%d.%m.%y",
    "%d.%m.%Y",
    "%d/%m/%Y",
    "%Y%m%d",
    "%Y-%m-%d",
    "%d-%m-%Y"
]

def parsear_fecha_multiple_conciliacion(valor):
    """Intenta convertir un valor a fecha probando varios formatos."""
    if pd.isna(valor):              # Verifica si el valor está vacío o es nulo
        return pd.NaT               # Si el valor es nulo (por ejemplo, está vacío o es NaN), devuelve pd.NaT (que significa "Not a Time", es decir, no hay fecha).
    valor = str(valor).strip()      # Convierte el valor a texto y elimina espacios en blanco al inicio y al final
    for formato in formatos_fecha:  # Intenta convertir el valor a fecha usando varios formatos
        try:
            return pd.to_datetime(valor, format=formato, errors="raise")
        except:
            continue
    return pd.NaT

# 3. Función genérica de transformación para conciliación bancaria
    
def transformar_extracto_conciliacion(df, fuente):
    
    if fuente not in reglas_fuente:
        raise ValueError(f"No hay reglas definidas para la fuente '{fuente}'")
    
    
    reglas = reglas_fuente.get(fuente)
    columnas = reglas['columnas']
  
    df_out_conciliacion = df.copy()

    
    # ✅ Columna: Tipo

    df_out_conciliacion['Tipo'] = fuente

    # ✅ Columna: Cuenta

    df_out_conciliacion['Cuenta'] = df.iloc[:, columnas['Cuenta']].astype(str).str.strip()

    # ✅ Columna: Fecha
    df_out_conciliacion['Fecha'] = (
        df.iloc[:, columnas['Fecha']]
        .astype(str)
        .str.strip()
        .apply(parsear_fecha_multiple_conciliacion)
    )    

    # ✅ Columna: Descripción

    df_out_conciliacion['Descripcion'] = df.iloc[:, columnas['Descripcion']].astype(str).str.strip()

    # ✅ Columna: importe

    df_out_conciliacion['Importe'] = pd.to_numeric(df.iloc[:, columnas['Importe']],errors="coerce").fillna(0)

    # ✅ Columna: Importe 2 (ajuste de signo para Auxiliar)
    if fuente == "Auxiliar":
        df_out_conciliacion['Importe 2'] = df_out_conciliacion['Importe'] * -1
    else:
        df_out_conciliacion['Importe 2'] = df_out_conciliacion['Importe']

    # ✅ Columnas opcionales

    opcionales_extracto = {
        'Asiento': lambda s: s.fillna("").astype(str).str.lstrip('0').str.upper(),
        'Nid': lambda s: s.fillna("").astype(str).str.lstrip('0').str.upper(),
        'IT Tercero': lambda s: s.fillna("").astype(str).str.upper(),
        'Tercero': lambda s: s.fillna("").astype(str).str.strip().str.upper(),
        'Responsable': lambda s: s.fillna("").astype(str).str.strip().str.upper(),
        'Fecha_hora modificacion': lambda s: s.fillna("").astype(str).str.strip(),
        'Ref 1': lambda s: s.fillna("").astype(str).str.upper(),
        'Ref 2': lambda s: s.fillna("").astype(str).str.lstrip('0').str.upper(),
    }

    for col, func in opcionales_extracto.items():
        if col in columnas:
            df_out_conciliacion[col] = func(df.iloc[:, columnas[col]])
        else:
            df_out_conciliacion[col] = ""


    df_final_conciliacion = df_out_conciliacion[['Tipo','Cuenta','Fecha', 'Asiento', 'Descripcion', 'Importe', 'Importe 2','Ref 1', 'Ref 2', 'Nid', 'IT Tercero', 'Tercero', 'Responsable', 'Fecha_hora modificacion']]
    return df_final_conciliacion

# 4. Reglas de cruce 

def conciliar_extractos(df_extracto, df_auxiliar):
    conciliados = []
    no_conciliados = []
    usados_aux = set()

    for i, row_ext in df_extracto.iterrows():
        encontrado = False

        for j, row_aux in df_auxiliar.iterrows():
            if j in usados_aux:
                continue

            # --- Reglas de cruce ---
            if (row_ext["Fecha"] == row_aux["Fecha"]) and \
               (row_ext["Descripcion"] == row_aux["Descripcion"]) and \
               (row_ext["Importe"] == row_aux["Importe"]):
                pass
            elif (row_ext["Fecha"] == row_aux["Fecha"]) and \
                 (row_ext["Importe"] == row_aux["Importe"]):
                pass
            elif abs((pd.to_datetime(row_ext["Fecha"], dayfirst=True) -
                      pd.to_datetime(row_aux["Fecha"], dayfirst=True)).days) <= 1 and \
                 (row_ext["Importe"] == row_aux["Importe"]):
                pass

            else:
                continue  # no match

            # ✅ Guardar match (conciliados)
            conciliados.append(row_ext.to_dict())
            conciliados.append(row_aux.to_dict())

            usados_aux.add(j)
            encontrado = True
            break

        if not encontrado:
            # Extracto sin match → guardarlo en bloque aparte
            no_conciliados.append(row_ext.to_dict())

    # Auxiliares no conciliados
    for j, row_aux in df_auxiliar.iterrows():
        if j not in usados_aux:
            no_conciliados.append(row_aux.to_dict())

    # Concatenar conciliados arriba + fila en blanco + no conciliados
    if conciliados and no_conciliados:
        conciliados.append({col: "" for col in df_extracto.columns})  # separador
    return pd.DataFrame(conciliados + no_conciliados)



# ----------------------------------- Interfaz Streamlit para conciliación bancaria -----------------------------------

# 🔹 Interfaz en Streamlit
st.title("Conciliación Bancaria")

col1, col2 = st.columns(2)

with col1:
    file_extracto = st.file_uploader("Sube archivo de Extracto", type=["xlsx"])
with col2:
    file_auxiliar = st.file_uploader("Sube archivo de Auxiliar", type=["xlsx"])

if file_extracto and file_auxiliar:
    df_ext_raw = pd.read_excel(file_extracto, header=None, skiprows=1)
    df_aux_raw = pd.read_excel(file_auxiliar, header=None, skiprows=1)

    df_ext = transformar_extracto_conciliacion(df_ext_raw, "Extracto")
    df_aux = transformar_extracto_conciliacion(df_aux_raw, "Auxiliar")

    st.subheader("Vista previa - Extracto")
    st.dataframe(df_ext.head(5))

    st.subheader("Vista previa - Auxiliar")
    st.dataframe(df_aux.head(5))

    # 🔹 Aplicar conciliación
    df_conciliado = conciliar_extractos(df_ext, df_aux)

    st.subheader("Resultado de la conciliación")
    st.dataframe(df_conciliado)

    # Descargar en Excel
    output_path = "conciliacion.xlsx"
    df_conciliado.to_excel(output_path, index=False, sheet_name="Conciliacion")
    

    # 🔹 Aplicar formato con openpyxl
    
    wb = load_workbook(output_path)
    ws = wb["Conciliacion"]

    # Buscar la columna "Importe" y "Tipo"
    col_importes = []
    for idx, cell in enumerate(ws[1], 1):  # Fila de encabezados
        if cell.value in ["Importe", "Importe 2"]:
            col_importes.append(idx)
        if cell.value == "Tipo":
            col_tipo = idx

    # Formato para columna "Importe"
    for col in col_importes:
        for row in ws.iter_rows(min_row=2, min_col=col, max_col=col):
            for cell in row:
                # 🔹 Formato: miles con ".", decimales con ",", negativos en rojo
                cell.number_format = '#,##0.00;[Red]-#,##0.00'

    # Color de fondo para filas con Tipo = Extracto
    fill_extracto = PatternFill(start_color="C5D9F1", end_color="C5D9F1", fill_type="solid")

    if col_tipo:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if row[col_tipo - 1].value == "Extracto":
                for cell in row:
                    cell.fill = fill_extracto


    wb.save("conciliacion.xlsx")

    with open("conciliacion.xlsx", "rb") as f:
        st.download_button("📥 Descargar conciliación en Excel", f, file_name="conciliacion.xlsx")





    